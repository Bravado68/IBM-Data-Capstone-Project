"""
Insert Interface Management - Arctic LNG2

Revise Script

@author: Reda.Hamidi

"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
startTime = datetime.now()

path ='C:\\Users\\saf167687\\OneDrive - Saipem\\Desktop\\18-Interface Management\\111-GBS3\\04-Interface_Management\\04-Script\\Python\\Library/'


#we shall store all the file names in this list
# filelist = []

# for root, dirs, files in os.walk(path):
# 	for file in files:
#         #append the file name to the list
# 		filelist.append(os.path.join(root,file))

#print all the file names
# for name in filelist:
#     print(name)
  
#sys.exit() 




# =============================================================================
# ############## Reading The input                         
# =============================================================================

GBS_Import=pd.read_excel('test.xlsx',sheet_name='Sheet1')
GBS_Master=pd.read_excel('test.xlsx',sheet_name='Revise')
GBS_Import=GBS_Import.drop(GBS_Import.index[0:3])
GBS_Import=GBS_Import.reset_index(drop=True)

GBS_Master=GBS_Master.drop(GBS_Import.index[0:3])
GBS_Master=GBS_Master.reset_index(drop=True)

GBS_Master = GBS_Master.fillna(0)
GBS_Import = GBS_Import.fillna(0)

# =============================================================================
# ############## Columns of Master and Import                       
# =============================================================================

GBS_Import.columns=["Mo Need Tag","Status","Date","Deleted","Parent_Tag","Insul","Temporary(Y/N)","Slab/Wall Tag","Watertight (Y/N)","Coordinate X","Coordinate Y","Coordinate Z","Diameter (mm)","Height (mm)","X length (mm)","Y length (mm)","Z length (mm)","Weight(Tons)","Reaction Loads X","Reaction Loads Y","Reaction Loads Z","Moments X","Moments Y","Moments Z","Tekla Tag",\
 "Remark","Pedestal_X","Pedestal_Y","Pedestal_X offset","Pedestal_Y offset","SAF","STR","MAE","CIV","Nav Ops","MOF","CIV Interface","Date interface","EP_Tag","E1","E2","E3","E4","E5","E6","Unique ID","Family","From Cell","PWBS","Rotz","Direction","Exist (y/n)","thickness"]
GBS_Master.columns=GBS_Import.columns


MO_Tag_import=GBS_Import.iloc[:,0].tolist()
Rev_import=GBS_Import.iloc[:,1].tolist()
Date_import=GBS_Import.iloc[:,2].tolist()
Dlt_import=GBS_Import.iloc[:,3].tolist()
Insert_Tag_import=GBS_Import.iloc[:,4].tolist()
INS_import=GBS_Import.iloc[:,5].tolist()
Tmp_Pmt_import=GBS_Import.iloc[:,6].tolist()
Location_Tag_import=GBS_Import.iloc[:,7].tolist()
WT_import=GBS_Import.iloc[:,8].tolist()
X_Tag_import=GBS_Import.iloc[:,9].tolist()
Y_Tag_import=GBS_Import.iloc[:,10].tolist()
Z_Tag_import=GBS_Import.iloc[:,11].tolist()
Geo_Diam_import=GBS_Import.iloc[:,12].tolist()
Thk_Tag_import=GBS_Import.iloc[:,13].tolist()
X_dim_import=GBS_Import.iloc[:,14].tolist()
Y_dim_import=GBS_Import.iloc[:,15].tolist()
Z_dim_import=GBS_Import.iloc[:,16].tolist()
Weight_import=GBS_Import.iloc[:,17].tolist()
RX_import=GBS_Import.iloc[:,18].tolist()
RY_import=GBS_Import.iloc[:,19].tolist()
RZ_import=GBS_Import.iloc[:,20].tolist()
MX_import=GBS_Import.iloc[:,21].tolist()
MY_import=GBS_Import.iloc[:,22].tolist()
MZ_import=GBS_Import.iloc[:,23].tolist()
Typ_import=GBS_Import.iloc[:,24].tolist()
Remark_import=GBS_Import.iloc[:,25].tolist()
Ped_X_import=GBS_Import.iloc[:,26].tolist()
Ped_Y_import=GBS_Import.iloc[:,27].tolist()
Ped_Offset_X_import=GBS_Import.iloc[:,28].tolist()
Ped_Offset_Y_import=GBS_Import.iloc[:,29].tolist()
Empty_import=GBS_Import.iloc[:,30].tolist()

MO_Tag_Master=GBS_Master.iloc[:,0].tolist()
Rev_Master=GBS_Master.iloc[:,1].tolist()
Date_Master=GBS_Master.iloc[:,2].tolist()
Dlt_Master=GBS_Master.iloc[:,3].tolist()
Insert_Tag_Master=GBS_Master.iloc[:,4].tolist()
INS_Master=GBS_Master.iloc[:,5].tolist()
Tmp_Pmt_Master=GBS_Master.iloc[:,6].tolist()
Location_Tag_Master=GBS_Master.iloc[:,7].tolist()
WT_Master=GBS_Master.iloc[:,8].tolist()
X_Tag_Master=GBS_Master.iloc[:,9].tolist()
Y_Tag_Master=GBS_Master.iloc[:,10].tolist()
Z_Tag_Master=GBS_Master.iloc[:,11].tolist()
Geo_Diam_Master=GBS_Master.iloc[:,12].tolist()
Thk_Tag_Master=GBS_Master.iloc[:,13].tolist()
X_dim_Master=GBS_Master.iloc[:,14].tolist()
Y_dim_Master=GBS_Master.iloc[:,15].tolist()
Z_dim_Master=GBS_Master.iloc[:,16].tolist()
Weight_Master=GBS_Master.iloc[:,17].tolist()
RX_Master=GBS_Master.iloc[:,18].tolist()
RY_Master=GBS_Master.iloc[:,19].tolist()
RZ_Master=GBS_Master.iloc[:,20].tolist()
MX_Master=GBS_Master.iloc[:,21].tolist()
MY_Master=GBS_Master.iloc[:,22].tolist()
MZ_Master=GBS_Master.iloc[:,23].tolist()
Typ_Master=GBS_Master.iloc[:,24].tolist()
Remark_Master=GBS_Master.iloc[:,25].tolist()
Ped_X_Master=GBS_Master.iloc[:,26].tolist()
Ped_Y_Master=GBS_Master.iloc[:,27].tolist()
Ped_Offset_X_Master=GBS_Master.iloc[:,28].tolist()
Ped_Offset_Y_Master=GBS_Master.iloc[:,29].tolist()
Empty_Master=GBS_Master.iloc[:,30].tolist()
Empty_Master=GBS_Master.iloc[:,31].tolist()
Empty_Master=GBS_Master.iloc[:,32].tolist()
Empty_Master=GBS_Master.iloc[:,33].tolist()
Empty_Master=GBS_Master.iloc[:,34].tolist()
Empty_Master=GBS_Master.iloc[:,35].tolist()
Empty_Master=GBS_Master.iloc[:,36].tolist()
Empty_Master=GBS_Master.iloc[:,37].tolist()
Tag_EP_Master=GBS_Master.iloc[:,38].tolist()
Empty_Master=GBS_Master.iloc[:,39].tolist()
Empty_Master=GBS_Master.iloc[:,40].tolist()
Empty_Master=GBS_Master.iloc[:,41].tolist()
Unique_ID_Master=GBS_Master.iloc[:,42].tolist()
Item_Tag_Master=GBS_Master.iloc[:,43].tolist()
Walls_ID_Master=GBS_Master.iloc[:,44].tolist()
PWBS_Master=GBS_Master.iloc[:,45].tolist()
Rot_z_Master=GBS_Master.iloc[:,46].tolist()
_Master=GBS_Master.iloc[:,47].tolist()
_Master=GBS_Master.iloc[:,48].tolist()
_Master=GBS_Master.iloc[:,49].tolist()
direc_x_y_Master=GBS_Master.iloc[:,50].tolist()
Exist_Master=GBS_Master.iloc[:,51].tolist()
Thk_Master=GBS_Master.iloc[:,52].tolist()

Master_Items=[MO_Tag_Master,Rev_Master,Date_Master,Dlt_Master,Insert_Tag_Master,INS_Master,Tmp_Pmt_Master,Location_Tag_Master,WT_Master,X_Tag_Master,Y_Tag_Master,Z_Tag_Master,Geo_Diam_Master,Thk_Tag_Master,X_dim_Master,Y_dim_Master,Z_dim_Master,Weight_Master,RX_Master,RY_Master,RZ_Master,MX_Master,MY_Master,MZ_Master,Typ_Master,Remark_Master,Ped_X_Master,Ped_Y_Master,Ped_Offset_X_Master,Ped_Offset_Y_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Tag_EP_Master,Empty_Master,Empty_Master,Empty_Master,Unique_ID_Master,Item_Tag_Master,Walls_ID_Master,PWBS_Master,Rot_z_Master,_Master,_Master,_Master,direc_x_y_Master,Exist_Master,Thk_Master]
Import_Items=[MO_Tag_import,Rev_import,Date_import,Dlt_import,Insert_Tag_import,INS_import,Tmp_Pmt_import,Location_Tag_import,WT_import,X_Tag_import,Y_Tag_import,Z_Tag_import,Geo_Diam_import,Thk_Tag_import,X_dim_import,Y_dim_import,Z_dim_import,Weight_import,RX_import,RY_import,RZ_import,MX_import,MY_import,MZ_import,Typ_import,Remark_import,Ped_X_import,Ped_Y_import,Ped_Offset_X_import,Ped_Offset_Y_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import]



n=len(MO_Tag_import)
k=len(MO_Tag_Master)
l=len(Import_Items)
m=len(Master_Items)



# =============================================================================
# ############## Finding duplicate in MO need                      
# =============================================================================


dup_MO=[]
liste=[]

for i in MO_Tag_import:
    if i not in liste:
        liste.append(i)
    else:
        dup_MO.append(i)
dup_MO

# =============================================================================
# ############## Make coordinates and dimensions as float                     
# =============================================================================


X_Tag_import=list(map(float, X_Tag_import))
Y_Tag_import=list(map(float, Y_Tag_import))
Z_Tag_import=list(map(float, Z_Tag_import))
Geo_Diam_import=list(map(float, Geo_Diam_import))
Thk_Tag_import=list(map(float, Thk_Tag_import))
X_dim_import=list(map(int, X_dim_import))
Y_dim_import=list(map(int, Y_dim_import))
Z_dim_import=list(map(int, Z_dim_import))
Ped_Offset_X_import=list(map(int, Ped_Offset_X_import))
Ped_Offset_Y_import=list(map(int, Ped_Offset_Y_import))

# =============================================================================
# ############## Proof reading script - Checking if all input are in correct format                         
# =============================================================================

for i in range(n):
    if type(X_Tag_import[i])!=float and type(X_Tag_import[i])!=int :
        print(MO_Tag_import[i], "has X as string not float")
    if type(Y_Tag_import[i])!=float and type(Y_Tag_import[i])!=int :
        print(MO_Tag_import[i], "has Y as string not float")
    if type(Z_Tag_import[i])!=float and type(Z_Tag_import[i])!=int :
        print(MO_Tag_import[i], "has Z as string not float")
    if type(Geo_Diam_import[i])!=float and type(Geo_Diam_import[i])!=int:
        print(MO_Tag_import[i], "has Geo_diam as string not float")
    if type(Thk_Tag_import[i])!=float and type(Thk_Tag_import[i])!=int :
        print(MO_Tag_import[i], "has thk as string not float")
        
    if type(X_dim_import[i])!=int:
        print(MO_Tag_import[i], "has X_dim as string not float")
    if type(Y_dim_import[i])!=int:
        print(MO_Tag_import[i], "has Y_dim as string not float")
    if type(Z_dim_import[i])!=int:
        print(MO_Tag_import[i], "has Z_dimas string not float")                
        
    if  Ped_X_import[i]!="merged" and Ped_X_import[i]!="no" :
        if type(Ped_X_import[i])!=int:
            print(MO_Tag_import[i], "Ped_X is not right")    
        
    if  Ped_Y_import[i]!="merged" and Ped_Y_import[i]!="no" :
        if type(Ped_Y_import[i])!=int:
            print(MO_Tag_import[i], "Ped_Y is not right")   
        
    if Ped_Y_import[i]!="no" :
        if type(Ped_Offset_X_import[i])!=int:
            print(MO_Tag_import[i], "Ped_Offset_X is not right")   
        
    if Ped_Y_import[i]!="no" :
        if type(Ped_Offset_Y_import[i])!=int :
            print(MO_Tag_import[i], "Ped_Offset_Y is not right")   


# =============================================================================
# ############## Reading PWBS and GRID text files                         
# =============================================================================

Data_Grid = pd.read_csv(path+'book.txt', delimiter = "\t")
X_Grid=Data_Grid.iloc[:,2].tolist()
Y_Grid=Data_Grid.iloc[:,7].tolist()
XName=Data_Grid.iloc[:,3].tolist()
YName=Data_Grid.iloc[:,8].tolist()

Data_PWBS = pd.read_csv(path+'PWBS.txt', delimiter = "\t")
Data_PWBS
Xmin=Data_PWBS.iloc[:,3].tolist()
Xmax=Data_PWBS.iloc[:,4].tolist()

Ymin=Data_PWBS.iloc[:,5].tolist()
Ymax=Data_PWBS.iloc[:,6].tolist()

Zmin=Data_PWBS.iloc[:,7].tolist()
Zmax=Data_PWBS.iloc[:,8].tolist()

PWBS_GRid=Data_PWBS.iloc[:,1].tolist()

Element_Grid=Data_PWBS.iloc[:,10].tolist()


# =============================================================================
# ############## Class Wall ID and PWBS                     
# =============================================================================

class Wall_Name():
    def __init__(self,location, x,y,z,Elem):
        
        self.location=location
        self.x=x
        self.y=y
        self.z=z
        self.Elem=Elem
        
        
    def Wall_ID(self):
        
        
        self.WallYName=0
        self.Wall_Name=[] 
        self.WallXName=0
        
        if int(self.x)<85000 or int(self.x)>415000:
            
            for j in range (len(X_Grid)):
                if int(self.x)>X_Grid[j] and int(self.x)<X_Grid[j+1]:
                    if XName[j]=='XX'and (XName[j + 1])=='1':
                        self.WallXName='XX01'
                        
                    elif XName[j]=='XX'and (XName[j + 1])=='XX':
                        self.WallXName='XXXX'
                        
                    
                    elif XName[j]=='40'and (XName[j + 1])=='XX':
                        self.WallXName='40XX'
                        
            
                
        else:
            
            for j in range (len(X_Grid)):
                            
                if int(self.x)>X_Grid[j] and int(self.x)<X_Grid[j+1]:
                    self.WallXName= XName[j] + XName[j + 1]
                    if int(float(XName[j]))< 10 and int(float((XName[j + 1]))) < 10: 
                        self.WallXName='0'+ XName[j] + '0'+ XName[j + 1]
                    elif int(float((XName[j]))) < 10 and int(float((XName[j + 1]))) >= 10:
                        self.WallXName='0'+ XName[j] + XName[j + 1]
                        
                    break
                    
        
                elif int(self.x)==X_Grid[j]:
                    self.WallXName = XName[j]+XName[j]
                    if int(XName[j])< 10: 
                        self.WallXName='0'+ XName[j] + '0'+ XName[j]
                    
                    break
            
                   
           
        for j in range (len(Y_Grid)):
            if int(self.y)>Y_Grid[j] and int(self.y)<Y_Grid[j+1]:
            
                self.WallYName = YName[j] + YName[j + 1]
                
                break
                
            elif int(self.y)==Y_Grid[j]:
                self.WallYName = YName[j]+YName[j]
                
                break
            
    
                
        self.Wall_ID=str(self.WallYName)+str(self.WallXName)
        self.Wall_Name.append(self.Wall_ID)
                
        return self.Wall_Name
    
    def PWBS_ID(self):
        self.PWBS_ID=[]
        self.PWBS_num=0
        if self.location=="PITSLAB":
            self.PWBS_ID=[3]
            return self.PWBS_ID
        
        elif self.location=="CORBEL":
            self.PWBS_ID=[5]
            return self.PWBS_ID
        
        elif self.location=="BUCKET":
            self.PWBS_ID=[2]
            return self.PWBS_ID
        
        elif self.location=="Pedestal":
            self.PWBS_ID=[0]
            return self.PWBS_ID  
        else:
            self.PWBS_num=99
            for j in range (len(Xmin)):
                if int(self.x)>Xmin[j] and int(self.x)<=Xmax[j] and int(self.y)>Ymin[j] and int(self.y)<=Ymax[j] and int(self.z)>Zmin[j] and int(self.z)<=Zmax[j] and self.Elem== Element_Grid[j]:
                    self.PWBS_num=PWBS_GRid[j]
                    break
       
            self.PWBS_ID.append(self.PWBS_num)        
            return self.PWBS_ID


# =============================================================================
# ############## Element "wall" or "slab"                       
# =============================================================================
            
Element=[]    
for i in range (k):
    if len(str(Location_Tag_Master[i]))<=2:
        Element.append('Wall')
    else:
        Element.append('Slab')   
        
Element_old=[]    
for i in range (n):
    if len(str(Location_Tag_import[i]))<=2:
        Element_old.append('Wall')
    else:
        Element_old.append('Slab')   


# =============================================================================
# ############## Determine Wall_ID and PWBS                         
# =============================================================================

Wall_ID_Tag=[]
PWBS=[]

for i in range(k):
    Location_Tag=Location_Tag_Master[i]
    X_Tag=X_Tag_Master[i]
    Y_Tag=Y_Tag_Master [i]
    Z_Tag=Z_Tag_Master [i]
    Elem=Element[i]
    
    A=Wall_Name(Location_Tag,X_Tag,Y_Tag,Z_Tag,Elem)
    Wall_ID_Tag.append(A.Wall_ID()[0])
    PWBS.append(A.PWBS_ID()[0])  
    
    
# =============================================================================
# ############## Correct format of date                    
# =============================================================================
    
Date_import=pd.to_datetime(Date_import, errors='coerce')
Date_import=Date_import.strftime('%m/%d/%Y').tolist()

# =============================================================================
# ############## List of modifications parameters                        
# =============================================================================

list_Rev_modified=[]
list_Date_modified=[]
list_Dlt_modified=[]
list_Descr_modified=[]
list_INS_modified=[]
list_Temporary=[]
list_Location_Tag_modified=[]
list_WT_modified=[]
list_X_modified=[]
list_Y_modified=[]
list_Z_modified=[]
list_diam_modified=[]
list_thk_modified=[]
list_Xdim_modified=[]
list_Ydim_modified=[]
list_Z_dim_Modified=[]
list_Weight_Modified=[]
list_RX_Modified=[]
list_RY_Modified=[]
list_RZ_Modified=[]
list_MX_Modified=[]
list_MY_Modified=[]
list_MZ_Modified=[]
list_Typ_Modified=[]
list_Remark_Modified=[]
list_Ped_X_Modified=[]
list_Ped_Y_Modified=[]
list_Ped_Offset_X_Modified=[]
list_Ped_Offset_Y_Modified=[]
list_WT_modified=[]

List_Modified=[list_Rev_modified,list_Date_modified,list_Dlt_modified,list_Descr_modified,list_INS_modified,list_Temporary,list_Location_Tag_modified,list_WT_modified,list_X_modified,list_Y_modified,list_Z_modified,list_diam_modified,list_thk_modified,list_Xdim_modified,list_Ydim_modified,list_Z_dim_Modified,list_Weight_Modified,list_RX_Modified,list_RY_Modified,list_RZ_Modified,list_MX_Modified,list_MY_Modified,list_MZ_Modified,list_Typ_Modified,list_Remark_Modified,list_Ped_X_Modified,list_Ped_Y_Modified,list_Ped_Offset_X_Modified,list_Ped_Offset_Y_Modified]



print("part1 .....before running script revise")
print (datetime.now() - startTime)        



# =============================================================================
# ############## Running the Revise script                        
# =============================================================================
Track=[]
Mo_Existing=[]
index_Existing=[]
Mo_deleted_change=[] 
Mo_deleted=[] 
for i in MO_Tag_Master:

    if i in MO_Tag_import:
        
        index_Old = MO_Tag_import.index(i)
        count_Mo=Track.count(i) 
        indices =[loc for loc, val in enumerate(MO_Tag_Master) if val == i]
        index_New =indices[count_Mo]
        Track.append(i)
        if Dlt_import[index_Old]!="DEL" : #and Exist_Master[index_New]=='TRUE':
            
            Location_old=Location_Tag_import[index_Old]
            X_old=X_Tag_import[index_Old]
            Y_old=Y_Tag_import[index_Old]
            Z_old=Z_Tag_import[index_Old]
            Elem_old=Element_old[index_Old]

            Location_new=Location_Tag_Master[index_New]
            X_new=X_Tag_Master[index_New]
            Y_new=Y_Tag_Master[index_New]
            Z_new=Z_Tag_Master[index_New]
            Elem_new=Element[index_New]

            Grid_Old=Wall_Name(Location_old,X_old,Y_old,Z_old,Elem_old)
            Grid_New=Wall_Name(Location_new,X_new,Y_new,Z_new,Elem_new)

            if Grid_Old.Wall_ID()==Grid_New.Wall_ID() and Grid_Old.PWBS_ID()==Grid_New.PWBS_ID():

                Exist_Master[index_New]="TRUE"
                Dlt_Master[index_New]="unset"
                Mo_Existing.append(i)
                index_Existing.append(index_New)

                if Rev_import[index_Old]=='Revised':
                    Rev_Master[index_New]+=1
                    
                
                for j in range(2,30):
            
                    if Import_Items[j][index_Old]!= Master_Items[j][index_New]:
                        List_Modified[j-1].append(i)
                        Master_Items[j][index_New]=Import_Items[j][index_Old]

            else:
                if Dlt_Master[index_New]=="DEL":
                    Exist_Master[index_New]="FALSE"
                else:
                    Dlt_Master[index_New]="DEL"
                    Mo_deleted_change.append(i)
            
        elif Dlt_import[index_Old]=="DEL":
            index_New = MO_Tag_Master.index(i)
            Exist_Master[index_New]="FALSE"
            Dlt_Master[index_New]="DEL"
        
    else:
        index_New = MO_Tag_Master.index(i)
        Exist_Master[index_New]="FALSE"
        Dlt_Master[index_New]="DEL"
        Mo_deleted.append(i)
        
print("part2....running revise script")
print (datetime.now() - startTime)



# =============================================================================
# ############## Making Dataframe of Import and Master input                      
# =============================================================================

GBS_Master=pd.DataFrame(list(zip(MO_Tag_Master,Rev_Master,Date_Master,Dlt_Master,Insert_Tag_Master,INS_Master,Tmp_Pmt_Master,Location_Tag_Master,WT_Master,X_Tag_Master,Y_Tag_Master,Z_Tag_Master,Geo_Diam_Master,Thk_Tag_Master,X_dim_Master,Y_dim_Master,Z_dim_Master,Weight_Master,RX_Master,RY_Master,RZ_Master,MX_Master,MY_Master,MZ_Master,Typ_Master,Remark_Master,Ped_X_Master,Ped_Y_Master,Ped_Offset_X_Master,Ped_Offset_Y_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Tag_EP_Master,Empty_Master,Empty_Master,Empty_Master,Unique_ID_Master,Item_Tag_Master,Walls_ID_Master,PWBS_Master,Rot_z_Master,_Master,_Master,_Master,direc_x_y_Master,Exist_Master,Thk_Master)))
GBS_Import=pd.DataFrame(list(zip(MO_Tag_import,Rev_import,Date_import,Dlt_import,Insert_Tag_import,INS_import,Tmp_Pmt_import,Location_Tag_import,WT_import,X_Tag_import,Y_Tag_import,Z_Tag_import,Geo_Diam_import,Thk_Tag_import,X_dim_import,Y_dim_import,Z_dim_import,Weight_import,RX_import,RY_import,RZ_import,MX_import,MY_import,MZ_import,Typ_import,Remark_import,Ped_X_import,Ped_Y_import,Ped_Offset_X_import,Ped_Offset_Y_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import,Empty_import)))

# =============================================================================
# ############## Filling 0 in cells without values                        
# =============================================================================

GBS_Master = GBS_Master.fillna(0)
GBS_Import= GBS_Import.fillna(0)

# =============================================================================
# ############## Correcting date format                        
# =============================================================================

GBS_Master[2]= pd.to_datetime(GBS_Master[2], errors='coerce')
GBS_Master[2]=GBS_Master[2].dt.strftime('%m/%d/%Y')


# =============================================================================
# ############## Creating New Dataframe with New MO need                       
# =============================================================================
Mo_New=[] 
df_NEW = GBS_Master[0:0]
for i in MO_Tag_import:
    index = MO_Tag_import.index(i)
    if Dlt_import[index]!="DEL":
        if i not in MO_Tag_Master:
            Mo_New.append(i)
            
MO_New_Final=Mo_deleted_change+Mo_New  

for i in MO_New_Final:
    df_NEW=pd.concat([df_NEW,GBS_Import[GBS_Import.iloc[:,0]==i]])
    
d=len(MO_Tag_Master)+len(MO_New_Final)
df_NEW.iloc[:, 51]='TRUE'
df_NEW.iloc[:,1]=1


df_NEW[2]= pd.to_datetime(df_NEW[2], errors='coerce')
df_NEW[2]=df_NEW[2].dt.strftime('%m/%d/%Y')
#df_NEW

print("part3.....creating new dataframe for new items")
print (datetime.now() - startTime)


# =============================================================================
# ############## Merging existing Master input and New Dataframe                      
# =============================================================================

GBS_Master=pd.concat([GBS_Master,df_NEW])


# =============================================================================
# ############## Converting to list the Import and Master dataframe                        
# =============================================================================

GBS_Master=GBS_Master.reset_index(drop=True)
#GBS_Master.columns=['MO_Tag_Master','Rev_Master','Date_Master','Dlt_Master','Insert_Tag_Master','INS_Master','Tmp_Pmt_Master','Location_Tag_Master','WT_Master','X_Tag_Master','Y_Tag_Master','Z_Tag_Master','Geo_Diam_Master','Thk_Tag_Master','X_dim_Master','Y_dim_Master','Z_dim_Master','Weight_Master','RX_Master','RY_Master','RZ_Master','MX_Master','MY_Master','MZ_Master','Typ_Master','Remark_Master','Ped_X_Master','Ped_Y_Master','Ped_Offset_X_Master','Ped_Offset_Y_Master','Empty_Master1','Empty_Master2','Empty_Master3','Empty_Master4','Empty_Master5','Empty_Master6','Empty_Master7','Empty_Master8','EP_Tag','Empty_Master10','Empty_Master11','Empty_Master12','Unique_ID_Master','Item_Tag_Master','Walls_ID_Master','PWBS_Master','Rot_z_Master','_Master','_Master1','_Master2','direc_x_y_Master','Exist_Master','Thk_Master']
GBS_Import.columns=["Mo Need Tag","Status","Date","Deleted","Parent_Tag","Insul","Temporary(Y/N)","Slab/Wall Tag","Watertight (Y/N)","Coordinate X","Coordinate Y","Coordinate Z","Diameter (mm)","Height (mm)","X length (mm)","Y length (mm)","Z length (mm)","Weight(Tons)","Reaction Loads X","Reaction Loads Y","Reaction Loads Z","Moments X","Moments Y","Moments Z","Tekla Tag",\
 "Remark","Pedestal_X","Pedestal_Y","Pedestal_X offset","Pedestal_Y offset","SAF","STR","MAE","CIV","Nav Ops","MOF","CIV Interface","Date interface","EP_Tag","E1","E2","E3","E4","E5","E6","Unique ID","Family","From Cell","PWBS","Rotz","Direction","Exist (y/n)","thickness"]
GBS_Master.columns=GBS_Import.columns



MO_Tag_Import=GBS_Import.iloc[:,0].tolist()
Rev_Import=GBS_Import.iloc[:,1].tolist()
Date_Import=GBS_Import.iloc[:,2].tolist()
Dlt_Import=GBS_Import.iloc[:,3].tolist()
Insert_Tag_Import=GBS_Import.iloc[:,4].tolist()
INS_Import=GBS_Import.iloc[:,5].tolist()
Tmp_Pmt_Import=GBS_Import.iloc[:,6].tolist()
Location_Tag_Import=GBS_Import.iloc[:,7].tolist()
WT_Import=GBS_Import.iloc[:,8].tolist()
X_Tag_Import=GBS_Import.iloc[:,9].tolist()
Y_Tag_Import=GBS_Import.iloc[:,10].tolist()
Z_Tag_Import=GBS_Import.iloc[:,11].tolist()
Geo_Diam_Import=GBS_Import.iloc[:,12].tolist()
Thk_Tag_Import=GBS_Import.iloc[:,13].tolist()
X_dim_import=GBS_Import.iloc[:,14].tolist()
Y_dim_Import=GBS_Import.iloc[:,15].tolist()
Z_dim_Import=GBS_Import.iloc[:,16].tolist()
Weight_Import=GBS_Import.iloc[:,17].tolist()
RX_Import=GBS_Import.iloc[:,18].tolist()
RY_Import=GBS_Import.iloc[:,19].tolist()
RZ_Import=GBS_Import.iloc[:,20].tolist()
MX_Import=GBS_Import.iloc[:,21].tolist()
MY_Import=GBS_Import.iloc[:,22].tolist()
MZ_Import=GBS_Import.iloc[:,23].tolist()
Typ_Import=GBS_Import.iloc[:,24].tolist()
Remark_Import=GBS_Import.iloc[:,25].tolist()
Nothing_Import=GBS_Import.iloc[:,26].tolist()
Ped_X_Import=GBS_Import.iloc[:,27].tolist()
Ped_Y_Import=GBS_Import.iloc[:,28].tolist()
Ped_Offset_X_Import=GBS_Import.iloc[:,29].tolist()
Ped_Offset_Y_Import=GBS_Import.iloc[:,30].tolist()


MO_Tag_Master=GBS_Master.iloc[:,0].tolist()
Rev_Master=GBS_Master.iloc[:,1].tolist()
Date_Master=GBS_Master.iloc[:,2].tolist()
Dlt_Master=GBS_Master.iloc[:,3].tolist()
Insert_Tag_Master=GBS_Master.iloc[:,4].tolist()
INS_Master=GBS_Master.iloc[:,5].tolist()
Tmp_Pmt_Master=GBS_Master.iloc[:,6].tolist()
Location_Tag_Master=GBS_Master.iloc[:,7].tolist()
WT_Master=GBS_Master.iloc[:,8].tolist()
X_Tag_Master=GBS_Master.iloc[:,9].tolist()
Y_Tag_Master=GBS_Master.iloc[:,10].tolist()
Z_Tag_Master=GBS_Master.iloc[:,11].tolist()
Geo_Diam_Master=GBS_Master.iloc[:,12].tolist()
Thk_Tag_Master=GBS_Master.iloc[:,13].tolist()
X_dim_Master=GBS_Master.iloc[:,14].tolist()
Y_dim_Master=GBS_Master.iloc[:,15].tolist()
Z_dim_Master=GBS_Master.iloc[:,16].tolist()
Weight_Master=GBS_Master.iloc[:,17].tolist()
RX_Master=GBS_Master.iloc[:,18].tolist()
RY_Master=GBS_Master.iloc[:,19].tolist()
RZ_Master=GBS_Master.iloc[:,20].tolist()
MX_Master=GBS_Master.iloc[:,21].tolist()
MY_Master=GBS_Master.iloc[:,22].tolist()
MZ_Master=GBS_Master.iloc[:,23].tolist()
Typ_Master=GBS_Master.iloc[:,24].tolist()
Remark_Master=GBS_Master.iloc[:,25].tolist()
Ped_X_Master=GBS_Master.iloc[:,26].tolist()
Ped_Y_Master=GBS_Master.iloc[:,27].tolist()
Ped_Offset_X_Master=GBS_Master.iloc[:,28].tolist()
Ped_Offset_Y_Master=GBS_Master.iloc[:,29].tolist()
Empty_Master=GBS_Master.iloc[:,30].tolist()
Empty_Master=GBS_Master.iloc[:,31].tolist()
Empty_Master=GBS_Master.iloc[:,32].tolist()
Empty_Master=GBS_Master.iloc[:,33].tolist()
Empty_Master=GBS_Master.iloc[:,34].tolist()
Empty_Master=GBS_Master.iloc[:,35].tolist()
Empty_Master=GBS_Master.iloc[:,36].tolist()
Empty_Master=GBS_Master.iloc[:,37].tolist()
Tag_EP_Master=GBS_Master.iloc[:,38].tolist()
Empty_Master=GBS_Master.iloc[:,39].tolist()
Empty_Master=GBS_Master.iloc[:,40].tolist()
Empty_Master=GBS_Master.iloc[:,41].tolist()
Unique_ID_Master=GBS_Master.iloc[:,42].tolist()
Item_Tag_Master=GBS_Master.iloc[:,43].tolist()
Walls_ID_Master=GBS_Master.iloc[:,44].tolist()
PWBS_Master=GBS_Master.iloc[:,45].tolist()
Rot_z_Master=GBS_Master.iloc[:,46].tolist()
_Master=GBS_Master.iloc[:,47].tolist()
_Master=GBS_Master.iloc[:,48].tolist()
_Master=GBS_Master.iloc[:,49].tolist()
direc_x_y_Master=GBS_Master.iloc[:,50].tolist()
Exist_Master=GBS_Master.iloc[:,51].tolist()
Thk_Master=GBS_Master.iloc[:,52].tolist()


# =============================================================================
# ############## Functions of highlight the modificated values                         
# =============================================================================


def style_Change_Date(x):
    if x[1]in list_Date_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_Descr(x):
    if x[1]in list_Descr_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2    

def style_Change_Ins(x):
    if x[1]in list_INS_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2

    

def style_temporary(x):
    if x[1]in list_Temporary:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_Location(x):
    if x[1]in list_Location_Tag_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2 
    
    
def style_Change_WT(x):
    if x[1]in list_WT_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2     

    
    
    
def style_Change_diam(x):
    if x[1]in list_diam_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_thk(x):
    if x[1]in list_thk_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2
    

def style_Change_X(x):
    if x[1]in list_X_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2
        
def style_Change_Y(x):
    if x[1]in list_Y_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2

    
def style_Change_Z(x):
    if x[1]in list_Z_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2
    


def style_Change_Xdim(x):
    if x[1]in list_Xdim_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2
    

def style_Change_Ydim(x):
    if x[1]in list_Ydim_modified:
        return ['background-color: gray']*2
    else:
        return ['']*2  


def style_Change_Zdim(x):
    if x[1]in list_Z_dim_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2

def style_Change_Weight(x):
    if x[1]in list_Weight_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2   

    
def style_Change_Typ(x):
    if x[1]in list_Typ_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2
    
    
def style_Change_Xped(x):
    if x[1]in list_Ped_X_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_Yped(x):
    if x[1]in list_Ped_Y_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_Ped_Offset_X(x):
    if x[1]in list_Ped_Offset_X_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2

def style_Change_Ped_Offset_Y(x):
    if x[1]in list_Ped_Offset_Y_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2


def style_Change_Remark(x):
    if x[1]in list_Remark_Modified:
        return ['background-color: gray']*2
    else:
        return ['']*2

def style_New(x):
    if x[0]in MO_New_Final :
        return ['background-color: yellow']*53
        
    else:
        return ['']*53  
    
def style_Delete(x):
    if x[3]=='DEL' :
        return ['background-color: red']*53
        
    else:
        return ['']*53 
    
Modified_items=[list_Rev_modified,list_Date_modified,list_Dlt_modified,list_Descr_modified,list_INS_modified,list_Temporary,list_Location_Tag_modified,list_WT_modified,list_X_modified,list_Y_modified,list_Z_modified,list_diam_modified,list_thk_modified,list_Xdim_modified,list_Ydim_modified,list_Z_dim_Modified,list_Weight_Modified,list_RX_Modified,list_RY_Modified,list_RZ_Modified,list_MX_Modified,list_MY_Modified,list_MZ_Modified,list_Typ_Modified,list_Remark_Modified,list_Ped_X_Modified,list_Ped_Y_Modified,list_Ped_Offset_X_Modified,list_Ped_Offset_Y_Modified,list_WT_modified]        

Modified_list=[]
for i in Modified_items:
    Modified_list=i+Modified_list
Modified_list = set(Modified_list)


# =============================================================================
# ############## New revision for the modified items                       
# =============================================================================
for i in Modified_list:
    index_New = MO_Tag_Master.index(i)
    Rev_Master[index_New]+=1


k=len(MO_Tag_Master)

count_TSP=[]
for i in range(k):
    if str(Typ_Master[i])[:3]=="TSP":
        
        d=count_TSP.count(Typ_Master[i])+1
        if d>99:
            Tag_EP_Master[i]=str(Typ_Master[i])+'-'+ str(d)
            
        elif d>9 and d<=99:
            Tag_EP_Master[i]=str(Typ_Master[i])+'-0'+str(d)   
        else:
            
            Tag_EP_Master[i]=str(Typ_Master[i])+'-00'+str(d) 
            
        count_TSP.append(Typ_Master[i])


# =============================================================================
# ############## Creating Master dataframe                       
# =============================================================================
Master_Items=[MO_Tag_Master,Rev_Master,Date_Master,Dlt_Master,Insert_Tag_Master,INS_Master,Tmp_Pmt_Master,Location_Tag_Master,WT_Master,X_Tag_Master,Y_Tag_Master,Z_Tag_Master,Geo_Diam_Master,Thk_Tag_Master,X_dim_Master,Y_dim_Master,Z_dim_Master,Weight_Master,RX_Master,RY_Master,RZ_Master,MX_Master,MY_Master,MZ_Master,Typ_Master,Remark_Master,Ped_X_Master,Ped_Y_Master,Ped_Offset_X_Master,Ped_Offset_Y_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Empty_Master,Tag_EP_Master,Empty_Master,Empty_Master,Empty_Master,Unique_ID_Master,Item_Tag_Master,Walls_ID_Master,PWBS_Master,Rot_z_Master,_Master,_Master,_Master,direc_x_y_Master,Exist_Master,Thk_Master]

for i in range (53):
    GBS_Master[GBS_Master.columns[i]]=Master_Items[i]


# =============================================================================
# ############## Correctng Insulation and WT cells                      
# =============================================================================

GBS_Master['Watertight (Y/N)']=GBS_Master['Watertight (Y/N)'].apply(lambda x: 'NA' if x==0 else x)
GBS_Master['Insul']=GBS_Master['Insul'].apply(lambda x: 'NA' if x==0 else x)


# =============================================================================
# ############## Creating blank instead of 0                 
# =============================================================================


GBS_Master['E1']=GBS_Master['E1'].replace(0, np.nan)
GBS_Master['E2']=GBS_Master['E2'].replace(0, np.nan)
GBS_Master['E3']=GBS_Master['E3'].replace(0, np.nan)
GBS_Master['E4']=GBS_Master['E4'].replace(0, np.nan)
GBS_Master['E5']=GBS_Master['E5'].replace(0, np.nan)
GBS_Master['E6']=GBS_Master['E6'].replace(0, np.nan)
GBS_Master['SAF']=GBS_Master['SAF'].replace(0, np.nan)
GBS_Master['STR']=GBS_Master['STR'].replace(0, np.nan)
GBS_Master['MAE']=GBS_Master['MAE'].replace(0, np.nan)
GBS_Master['CIV']=GBS_Master['CIV'].replace(0, np.nan)
GBS_Master['CIV Interface']=GBS_Master['CIV Interface'].replace(0, np.nan)
GBS_Master['Nav Ops']=GBS_Master['Nav Ops'].replace(0, np.nan)
GBS_Master['MOF']=GBS_Master['MOF'].replace(0, np.nan)
GBS_Master['Date interface']=GBS_Master['Date interface'].replace(0, np.nan)
GBS_Master['EP_Tag']=GBS_Master['EP_Tag'].replace(0, np.nan)





# =============================================================================
# ############## Deleting Revise,Export and Extra sheets                         
# =============================================================================

from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
if 'Revise' in wb.sheetnames:
    wb.remove(wb['Revise'])
wb.save('test.xlsx')
if 'Export' in wb.sheetnames:
    wb.remove(wb['Export'])
wb.save('test.xlsx')
if 'Extra' in wb.sheetnames:
    wb.remove(wb['Extra'])
wb.save('test.xlsx')


print('part4 ......highlight begins')
print (datetime.now() - startTime)





# =============================================================================
# ############### Creating Revise sheet without removing other sheets        
# ############### Highlighting modified and new values                 
# =============================================================================
#.apply(style_Change_Ins, subset=['INS_Master','Mo Need Tag'], axis=1)\
#.apply(style_Change_Pmt, subset=['Tmp_Pmt_Master','Mo Need Tag'], axis=1)\



#writing to excel without removing other sheets


#style_Change_Date, subset=['Date','Mo Need Tag'], axis=1)

#writing to excel without removing other sheets
with pd.ExcelWriter('test.xlsx', engine="openpyxl", mode="a") as writer:
    GBS_Master.style.apply(style_temporary, subset=['Temporary(Y/N)','Mo Need Tag'], axis=1)\
    .apply(style_Change_Ins, subset=['Insul','Mo Need Tag'], axis=1)\
    .apply(style_Change_WT, subset=['Watertight (Y/N)','Mo Need Tag'], axis=1)\
    .apply(style_Change_X, subset=['Coordinate X','Mo Need Tag'], axis=1)\
    .apply(style_Change_Y, subset=['Coordinate Y','Mo Need Tag'], axis=1)\
    .apply(style_Change_Z, subset=['Coordinate Z','Mo Need Tag'], axis=1)\
    .apply(style_Change_Descr, subset=['Parent_Tag','Mo Need Tag'], axis=1)\
    .apply(style_Change_diam, subset=['Diameter (mm)','Mo Need Tag'], axis=1)\
    .apply(style_Change_thk, subset=['Height (mm)','Mo Need Tag'], axis=1)\
    .apply(style_Change_Xdim, subset=['X length (mm)','Mo Need Tag'], axis=1)\
    .apply(style_Change_Ydim, subset=['Y length (mm)','Mo Need Tag'], axis=1)\
    .apply(style_Change_Zdim, subset=['Z length (mm)','Mo Need Tag'], axis=1)\
    .apply(style_Change_Xped, subset=['Pedestal_X','Mo Need Tag'], axis=1)\
    .apply(style_Change_Yped, subset=['Pedestal_Y','Mo Need Tag'], axis=1)\
    .apply(style_Change_Ped_Offset_X, subset=['Pedestal_X offset','Mo Need Tag'], axis=1)\
    .apply(style_Change_Ped_Offset_Y, subset=['Pedestal_Y offset','Mo Need Tag'], axis=1)\
    .apply(style_Change_Location, subset=['Slab/Wall Tag','Mo Need Tag'], axis=1)\
    .apply(style_Change_Weight, subset=['Weight(Tons)','Mo Need Tag'], axis=1)\
    .apply(style_New, axis=1)\
    .apply(style_Delete, axis=1)\
    .apply(style_Change_Remark, subset=['Remark','Mo Need Tag'], axis=1).to_excel(writer, sheet_name="Revise",index=False,startrow=3 , startcol=0)



print("part5............Script Revise finished")
print (datetime.now() - startTime)



        

     